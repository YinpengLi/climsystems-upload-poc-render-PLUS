import os, json, csv, datetime
from typing import Dict, Any
import pandas as pd
from openpyxl import load_workbook
from .db import connect

LAT_KEYS = ["lat","latitude","y"]
LON_KEYS = ["lon","lng","longitude","x"]

def _now():
    return datetime.datetime.utcnow().isoformat() + "Z"

def detect_columns(file_path: str) -> Dict[str, Any]:
    ext = os.path.splitext(file_path)[1].lower()
    cols, sample = [], []
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.DictReader(f)
            cols = reader.fieldnames or []
            for i, row in enumerate(reader):
                sample.append(row)
                if i >= 4: break
    elif ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=1, max_row=6, values_only=True)
        header = next(rows)
        cols = [str(c).strip() if c is not None else "" for c in header]
        for r in rows:
            d = {cols[i]: r[i] for i in range(min(len(cols), len(r)))}
            sample.append(d)
        wb.close()
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    lower = [c.lower() for c in cols]
    def best_match(keys):
        for k in keys:
            for i, c in enumerate(lower):
                if c == k or c.endswith(k) or k in c:
                    return cols[i]
        return None

    guess = {
        "lat_col": best_match(LAT_KEYS),
        "lon_col": best_match(LON_KEYS),
        "asset_id_col": best_match(["asset_id","asset","id","site","name"]),
        "label_col": best_match(["label","name","site","address","asset_name"]),
        "theme_col": best_match(["theme"]),
        "indicator_col": best_match(["indicator","metric","variable"]),
        "value_col": best_match(["value","score","val"]),
        "year_col": best_match(["year","yr"]),
        "scenario_col": best_match(["scenario","ssp","rcp"]),
        "units_col": best_match(["units","unit"]),
    }
    return {"columns": cols, "guess": guess, "sample": sample[:5]}

def ingest_to_sqlite(dataset_id: str, file_path: str, mapping: Dict[str, Any]) -> Dict[str, Any]:
    ext = os.path.splitext(file_path)[1].lower()
    con = connect()
    cur = con.cursor()
    cur.execute("DELETE FROM facts WHERE dataset_id=?", (dataset_id,))
    cur.execute("DELETE FROM assets WHERE dataset_id=?", (dataset_id,))
    con.commit()

    lat_col = mapping.get("lat_col")
    lon_col = mapping.get("lon_col")
    asset_id_col = mapping.get("asset_id_col") or "asset_id"
    label_col = mapping.get("label_col") or asset_id_col
    theme_col = mapping.get("theme_col") or "theme"
    indicator_col = mapping.get("indicator_col") or "indicator"
    value_col = mapping.get("value_col") or "value"
    year_col = mapping.get("year_col")
    scenario_col = mapping.get("scenario_col")
    units_col = mapping.get("units_col")

    asset_seen = {}
    row_count = 0
    themes, years, scenarios, indicators = set(), set(), set(), set()

    def normalize_row(row):
        asset_id = str(row.get(asset_id_col) or row.get("asset_id") or "").strip() or "UNKNOWN"
        lat = row.get(lat_col) if lat_col else row.get("latitude")
        lon = row.get(lon_col) if lon_col else row.get("longitude")
        try: lat = float(lat) if lat not in [None,""] else None
        except: lat = None
        try: lon = float(lon) if lon not in [None,""] else None
        except: lon = None

        theme = row.get(theme_col)
        indicator = row.get(indicator_col)
        value = row.get(value_col)

        theme = str(theme).strip() if theme not in [None,""] else None
        indicator = str(indicator).strip() if indicator not in [None,""] else None
        try: value = float(value) if value not in [None,""] else None
        except: value = None

        year = None
        if year_col and row.get(year_col) not in [None,""]:
            try: year = int(float(row.get(year_col)))
            except: year = None

        scenario = None
        if scenario_col and row.get(scenario_col) not in [None,""]:
            scenario = str(row.get(scenario_col)).strip()

        units = None
        if units_col and row.get(units_col) not in [None,""]:
            units = str(row.get(units_col)).strip()

        label = row.get(label_col)
        label = str(label).strip() if label not in [None,""] else asset_id

        extras = {}
        for k,v in row.items():
            if k in [asset_id_col,label_col,lat_col,lon_col,theme_col,indicator_col,value_col,year_col,scenario_col,units_col]:
                continue
            if k is None or str(k).strip()=="":
                continue
            extras[str(k)] = v

        return asset_id, lat, lon, label, year, scenario, theme, indicator, value, units, (json.dumps(extras) if extras else None)

    def upsert_asset(asset_id, lat, lon, label):
        if asset_id in asset_seen: return
        if lat is None or lon is None: return
        asset_seen[asset_id]=True
        cur.execute("INSERT INTO assets(dataset_id, asset_id, latitude, longitude, label) VALUES (?,?,?,?,?)",
                    (dataset_id, asset_id, lat, lon, label))

    def insert_fact(vals):
        nonlocal row_count
        cur.execute("INSERT INTO facts(dataset_id, asset_id, latitude, longitude, year, scenario, theme, indicator, value, units, extra_json) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (dataset_id,)+vals[:10]+(vals[10],))
        row_count += 1

    if ext == ".csv":
        for chunk in pd.read_csv(file_path, chunksize=100000):
            for _, r in chunk.iterrows():
                vals = normalize_row(r.to_dict())
                asset_id, lat, lon, label, year, scenario, theme, indicator, value, units, extra_json = vals
                if theme: themes.add(theme)
                if indicator: indicators.add(indicator)
                if year is not None: years.add(year)
                if scenario: scenarios.add(scenario)
                upsert_asset(asset_id, lat, lon, label)
                insert_fact((asset_id, lat, lon, year, scenario, theme, indicator, value, units, extra_json))
            con.commit()
    elif ext in [".xlsx",".xlsm",".xltx",".xltm"]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        cols = [str(c).strip() if c is not None else "" for c in header]
        batch = []
        for r in rows:
            d = {cols[i]: r[i] for i in range(min(len(cols), len(r)))}
            batch.append(d)
            if len(batch) >= 5000:
                for rr in batch:
                    vals = normalize_row(rr)
                    asset_id, lat, lon, label, year, scenario, theme, indicator, value, units, extra_json = vals
                    if theme: themes.add(theme)
                    if indicator: indicators.add(indicator)
                    if year is not None: years.add(year)
                    if scenario: scenarios.add(scenario)
                    upsert_asset(asset_id, lat, lon, label)
                    insert_fact((asset_id, lat, lon, year, scenario, theme, indicator, value, units, extra_json))
                con.commit()
                batch=[]
        if batch:
            for rr in batch:
                vals = normalize_row(rr)
                asset_id, lat, lon, label, year, scenario, theme, indicator, value, units, extra_json = vals
                if theme: themes.add(theme)
                if indicator: indicators.add(indicator)
                if year is not None: years.add(year)
                if scenario: scenarios.add(scenario)
                upsert_asset(asset_id, lat, lon, label)
                insert_fact((asset_id, lat, lon, year, scenario, theme, indicator, value, units, extra_json))
            con.commit()
        wb.close()
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    con.close()
    return {
        "row_count": row_count,
        "asset_count": len(asset_seen),
        "themes": sorted(themes),
        "years": sorted(years),
        "scenarios": sorted(scenarios),
        "indicators": sorted(indicators),
        "generated_at": _now()
    }
